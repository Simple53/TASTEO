#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
局域网点单系统 - Flask 主应用
支持用户点餐、商家管理、数据导出
"""

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime
import json
import os
import sqlite3
from contextlib import contextmanager
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import socket

app = Flask(__name__)
CORS(app)

# 配置
DATABASE = 'database.db'
app.config['JSON_AS_ASCII'] = False

# ==================== 数据库操作 ====================

@contextmanager
def get_db():
    """数据库连接上下文管理器"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()

def init_db():
    """初始化数据库"""
    with get_db() as conn:
        cursor = conn.cursor()
        
        # 创建菜单表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS menu (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                price REAL NOT NULL,
                image TEXT DEFAULT 'default.jpg',
                category TEXT DEFAULT '主食',
                available INTEGER DEFAULT 1
            )
        ''')
        
        # 创建订单表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                table_number TEXT NOT NULL,
                customer_name TEXT NOT NULL,
                items TEXT NOT NULL,
                total_price REAL NOT NULL,
                status TEXT DEFAULT 'pending',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                completed_at TIMESTAMP
            )
        ''')
        
        # 检查是否需要插入示例数据
        cursor.execute('SELECT COUNT(*) FROM menu')
        if cursor.fetchone()[0] == 0:
            # 插入示例菜单数据
            sample_menu = [
                ('宫保鸡丁', 38.0, 'default.jpg', '热菜'),
                ('鱼香肉丝', 35.0, 'default.jpg', '热菜'),
                ('麻婆豆腐', 28.0, 'default.jpg', '热菜'),
                ('糖醋里脊', 42.0, 'default.jpg', '热菜'),
                ('清蒸鲈鱼', 68.0, 'default.jpg', '海鲜'),
                ('红烧排骨', 48.0, 'default.jpg', '热菜'),
                ('西红柿炒蛋', 22.0, 'default.jpg', '家常菜'),
                ('酸辣土豆丝', 18.0, 'default.jpg', '凉菜'),
                ('米饭', 3.0, 'default.jpg', '主食'),
                ('紫菜蛋花汤', 15.0, 'default.jpg', '汤类'),
                ('可乐', 8.0, 'default.jpg', '饮料'),
                ('雪碧', 8.0, 'default.jpg', '饮料'),
            ]
            cursor.executemany(
                'INSERT INTO menu (name, price, image, category) VALUES (?, ?, ?, ?)',
                sample_menu
            )

# ==================== 路由 ====================

@app.route('/')
def index():
    """首页重定向到用户端"""
    return render_template('customer.html')

@app.route('/customer')
def customer():
    """用户点餐页面"""
    return render_template('customer.html')

@app.route('/merchant')
def merchant():
    """商家管理页面"""
    return render_template('merchant.html')

# ==================== API 接口 ====================

@app.route('/api/menu', methods=['GET'])
def get_menu():
    """获取菜单列表"""
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM menu WHERE available = 1 ORDER BY category, id')
            menu_items = cursor.fetchall()
            
            menu_list = []
            for item in menu_items:
                menu_list.append({
                    'id': item['id'],
                    'name': item['name'],
                    'price': item['price'],
                    'image': f"/static/images/{item['image']}",
                    'category': item['category']
                })
            
            return jsonify(menu_list)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/order', methods=['POST'])
def create_order():
    """创建订单"""
    try:
        data = request.json
        table_number = data.get('table_number', '')
        customer_name = data.get('customer_name', '')
        items = data.get('items', [])
        total_price = data.get('total_price', 0)
        
        if not table_number or not customer_name or not items:
            return jsonify({'error': '请填写完整信息'}), 400
        
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute(
                'INSERT INTO orders (table_number, customer_name, items, total_price) VALUES (?, ?, ?, ?)',
                (table_number, customer_name, json.dumps(items, ensure_ascii=False), total_price)
            )
            order_id = cursor.lastrowid
        
        return jsonify({'success': True, 'order_id': order_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/orders', methods=['GET'])
def get_orders():
    """获取所有订单"""
    try:
        status_filter = request.args.get('status', None)
        
        with get_db() as conn:
            cursor = conn.cursor()
            if status_filter:
                cursor.execute(
                    'SELECT * FROM orders WHERE status = ? ORDER BY created_at DESC',
                    (status_filter,)
                )
            else:
                cursor.execute('SELECT * FROM orders ORDER BY created_at DESC')
            
            orders = cursor.fetchall()
            
            order_list = []
            for order in orders:
                order_list.append({
                    'id': order['id'],
                    'table_number': order['table_number'],
                    'customer_name': order['customer_name'],
                    'items': order['items'],
                    'total_price': order['total_price'],
                    'status': order['status'],
                    'created_at': order['created_at'],
                    'completed_at': order['completed_at']
                })
            
            return jsonify(order_list)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/order/<int:order_id>/complete', methods=['PUT'])
def complete_order(order_id):
    """完成订单"""
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute(
                'UPDATE orders SET status = ?, completed_at = ? WHERE id = ?',
                ('completed', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), order_id)
            )
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    """导出 Excel"""
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM orders ORDER BY created_at DESC')
            orders = cursor.fetchall()
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "订单数据"
        
        # 设置表头
        headers = ['订单号', '桌号', '顾客姓名', '菜品详情', '总价', '状态', '下单时间', '完成时间']
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 填充数据
        for order in orders:
            items = json.loads(order['items'])
            items_str = ', '.join([f"{item['name']}x{item['quantity']}" for item in items])
            
            ws.append([
                order['id'],
                order['table_number'],
                order['customer_name'],
                items_str,
                order['total_price'],
                '已完成' if order['status'] == 'completed' else '待处理',
                order['created_at'],
                order['completed_at'] or ''
            ])
        
        # 调整列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        
        # 保存文件
        filename = f'orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        filepath = os.path.join('exports', filename)
        os.makedirs('exports', exist_ok=True)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/export/csv', methods=['GET'])
def export_csv():
    """导出 CSV"""
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM orders ORDER BY created_at DESC')
            orders = cursor.fetchall()
        
        # 准备数据
        data = []
        for order in orders:
            items = json.loads(order['items'])
            items_str = ', '.join([f"{item['name']}x{item['quantity']}" for item in items])
            
            data.append({
                '订单号': order['id'],
                '桌号': order['table_number'],
                '顾客姓名': order['customer_name'],
                '菜品详情': items_str,
                '总价': order['total_price'],
                '状态': '已完成' if order['status'] == 'completed' else '待处理',
                '下单时间': order['created_at'],
                '完成时间': order['completed_at'] or ''
            })
        
        # 创建 DataFrame
        df = pd.DataFrame(data)
        
        # 保存文件
        filename = f'orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        filepath = os.path.join('exports', filename)
        os.makedirs('exports', exist_ok=True)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== 工具函数 ====================

def get_local_ip():
    """获取本机局域网 IP"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return '127.0.0.1'

# ==================== 启动应用 ====================

if __name__ == '__main__':
    # 初始化数据库
    init_db()
    
    # 获取本机 IP
    local_ip = get_local_ip()
    
    print("=" * 60)
    print("🍜 点点鲜 - 局域网点单系统")
    print("=" * 60)
    print(f"📱 用户点餐页面: http://{local_ip}:5000/customer")
    print(f"💼 商家管理页面: http://{local_ip}:5000/merchant")
    print(f"🌐 本机访问: http://127.0.0.1:5000")
    print("=" * 60)
    print("提示: 局域网内其他设备可通过上述 IP 地址访问")
    print("按 Ctrl+C 停止服务")
    print("=" * 60)
    
    # 启动 Flask 应用
    app.run(host='0.0.0.0', port=5000, debug=False)
